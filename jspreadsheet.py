"""
Author: Justin Francisco
Date Written: 29/03/2023
Last Modified By: Justin Francisco

Description: @TCPwithFeels
Script to read, write or do anything with spreadsheets/csv
Dependencies: openpyxl
"""

import openpyxl
import jchecker
from datetime import date

today = date.today()
date_format = today.strftime("%d/%m/%Y")

def get_list_mac_hosts(SPREADSHEET="WAP info for SmartZone.xlsx", WORKSHEET="H510"):
    
    # Load the Spreadsheet
    wb = openpyxl.load_workbook(SPREADSHEET)

    # Find Correct Worksheet
    ws = wb[WORKSHEET]

    max_row = ws.max_row + 1
    mac_hostname_waplist = []

    # Return list of dictionaries
    for iteration in range(2,max_row):

        mac = ws["B{}".format(iteration)].value
        mac = jchecker.check_ruckus_mac(mac)
        DICT = {
            "name" : ws["A{}".format(iteration)].value,
            "location" : ws["C{}".format(iteration)].value,
            "mac"  : mac,
            "zoneId": "self.zone_id",
            "description": ws["D{}".format(iteration)].value
            #"apGroupId": "self.group_id",
            #"model": "Ruckus R510"
        }

        mac_hostname_waplist.append(DICT)
        
    return mac_hostname_waplist