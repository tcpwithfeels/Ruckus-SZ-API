"""
Author: Justin Francisco
Date Written: 23/03/2023
Last Modified By: Justin Francisco

Date Last Modified: YYYY/MM/DD
Date Last Tested: YYYY/MM/DD

Result: Pass / Fail

Description: Interaction with Ruckus SZ 100 API to Mass Add Ruckus 510 WAPs
Dependencies: requests json
Usage: `python3 ruckus_AP_add.py`

<description>

"""

import requests
import cprint
import re
import openpyxl
import json
from getpass import getpass

"""
API Documentation
https://docs.ruckuswireless.com/smartzone/6.1.0/sz100-public-api-reference-guide-610.html

-------

Common Request Header
The following parameters are required in the HTTP headers of all API requests.

Parameter:      Content-Type	
Value:          “application/json;charset=UTF-8”

-------

Common Request URI Parameters 
The following parameters are required in the Request URI Parameters of all API requests (except for the logon API).

Parameter:	    ServiceTicket	
Value:          {serviceTicket}

serviceTicket is returned as the following parameter in the response payload of the Service Ticket Logon API.
"""

# RUCKUS MAC FORMAT E.G.
# 341593017F00
def check_ruckus_mac(mac_address):
    # if re.match(r"^([0-9a-fA-F]{2}:){5}([0-9a-fA-F][0-9a-fA-F])$",mac_address):
    if re.match(r"^(341593)([0-9a-fA-F]{6})$",mac_address):
        print("""
              Valid MAC Address
              """)
        return True
    else:
        print("You inputted: {} \n This is not a valid MAC Address".format(mac_address))
        print("""        

Please Check the Following:
- It should be a 12 HEXADECIMAL Value
- Ruckus MAC Addresses do not contain : (colons)
- IOU of Ruckuz 510s starts with the following 341593

              """)
        return False

def host_name_checker(hostname):
    if re.match(r"^(B([0-9]){2}-L([0-9]){2}-R([0-9]){2})$",hostname):
        print("""
              Valid Hostname Address
              """)
        return True
    else:
        print("You inputted: {} \n This is not a valid Hostname".format(hostname))
        print("""        
Please Check the Following:
- Have two dash (-) separators
- 1st portion should start with B, BXX (XX = Building Number)
- 2nd portion should start with L, LXX (XX = Floor Number e.g. 1F)
- 2nd portion should start with R, RXX 
- FORMAT should be of BXX-LXX-RXX e.g. B98-L76-R54
              """)
        return False
    
class ruckus_SZ_API:
    # 27/03/2023
    def __init__(self, host: str):

        self.prefix_pattern = "https://{}:8443/wsg/api/public".format(host)

    def retrieve_api_version(self):

        # Get the API Version
        # This will be included in most requests in the URL portion

        url = "{}/apiInfo".format(self.prefix_pattern)
        self.api_version = requests.get(url)
        return self.api_version
    
    # -----------------------
    # Service Ticket - Logon
    # -----------------------

    def service_ticket_logon(self, api_version, username, password):

        # Use this API command to log on to the controller and acquire a valid service ticket.

        url = "{}/{}/serviceTicket".format(self.prefix_pattern, self.api_version)
        request_body = {
            "username": username,
            "password": password
        }

        try:
            r = requests.post(request_body, data=request_body, verify=False)
        except requests.exceptions.ConnectionError as e:
            cprint("Unable to establish session to Ruckus Smart Zone:\n  ", 'red', True)
            cprint(e,'yellow')
            exit()
        
        self.required_headers = {
            "Content-Type":  "application/json;charset=UTF-8",
            "serviceTicket": r["serviceTicket"]
        }

    # -----------------------
    # Session ID - Logon
    # -----------------------

    def session_id_logon(self, username, password):
        # DEPRICATED unless you're using old API Version
        # Ruckus Logon
        # Use this API command to log on to the controller and acquire a valid logon session.

        url = "{}/{}/sessions".format(self.prefix_pattern, self.api_version)

        # Request Body in POST
        # Required [username, password]
        request_body = {
            "username": username,
            "password": password
        }
        self.service_ticket_value = request_body["serviceTicket"]

        try:
            r = requests.post(url, data=request_body, verify=False)
        except requests.exceptions.ConnectionError as e:
            cprint("Unable to establish session to Ruckus Smart Zone:\n  ", 'red', True)
            cprint(e,'yellow')
            exit()
   
    # ------------------------------------------
    # Ruckus Wireless AP Zone - Retrieve List
    # Retrieve Zone ID from List
    # ------------------------------------------

    def retrieve_zone_id(self, zone_name):

        url = "{}/{}/rkszones".format(self.prefix_pattern, self.api_version)
        self.zone_name = zone_name

        """
        {
            "totalCount": 2,
            "hasMore": false,
            "firstIndex": 0,
            "list": [
                {
                "id": "zoneUUID",
                "name": "zoneName"
                },
                {
                "id": "zoneUUID2",
                "name": "zoneName2"
                }
            ]
        }
        """

        r = requests.get(url, headers=self.required_headers, verify=False)
        list_of_zones = r["list"]
        print("List of Zones Here:\n{}".format(list_of_zones))
        for zones in list_of_zones:
            if zones["name"] == zone_name:
                self.zone_id = zones["id"]
                print("""
                ZONE Name {} FOUND\nID Found: {}
                """.format(zone_name, self.zone_id)
                )

    # ------------------------------------------
    # AP Group - Retrieve List
    # Retrieve Group ID from List
    # ------------------------------------------

    def retrieve_group_id(self, group_name):

        url = "{}/{}/rkszones/{}/apgroups".format(self.prefix_pattern, self.api_version, self.zone_id)
        self.group_name = group_name

        """
        {
            "totalCount": 2,
            "hasMore": false,
            "firstIndex": 0,
            "list": [
                {
                "id": "apGroupUUID",
                "name": "apGroupName"
                },
                {
                "id": "apGroupUUID2",
                "name": "apGroupName2"
                }
            ]   
        }
        """

        r = requests.get(url, headers=self.required_headers, verify=False)
        list_of_groups = r["list"]
        print("List of Groups Here:\n{}".format(list_of_groups))
        for groups in list_of_groups:
            if groups["name"] == group_name:
                self.group_id = groups["id"]
                print("""
                Group Name {} FOUND\nID Found: {}
                """.format(group_name, self.group_id)
                )

    # ------------------------------------------
    # AP Group - Create WAP
    # CREATE AP - Specify MAC, Zone and Group
    # ------------------------------------------

    def create_ruckus_ap(self, host_name, mac):
        # "required" : [ "mac", "zoneId" ]
        url = "{}/{}/aps".format(self.prefix_pattern, self.api_version)

        if check_ruckus_mac(mac) and host_name_checker(host_name):
            
            AP_Info = {
                "mac": mac,
                "zoneId": self.zone_id,
                "apGroupId": self.group_id,
                "model": "Ruckus R510",
                "name": host_name
            }

            r = requests.post(url, data=AP_Info, headers=self.required_headers, verify=False)

        return r.status_code
    
    # ------------------------------------------
    # AP Configuration - Create WAP
    # CREATE AP - Specify MAC, Zone and Group
    # ------------------------------------------

    def verify_ruckus_ap(self, mac):
        # "required" : [ "mac", "zoneId" ]
        url = "{}/{}/aps/{}".format(self.prefix_pattern, self.api_version, mac)

        r = requests.post(url, headers=self.required_headers, verify=False)
        json_to_dictionary = json.dumps(r)


        return r.status_code
    
    def get_list_mac_hosts(self, SPREADSHEET):
        # Load the Spreadsheet

        wb = openpyxl.load_workbook(SPREADSHEET)

        ############################
        ws = wb["<VALUE>"]
        ############################

        max_row = ws.max_row + 1
        mac_hostname_waplist = []

        # Return list of dictionaries
        for iteration in range(2,max_row):
            
                DICT = {
                    "name" : ws["A{}".format(iteration)].value,
                    "mac"  : ws["B{}".format(iteration)].value,
                    "zoneId": "self.zone_id",
                    "apGroupId": "self.group_id"
                    "model": "Ruckus R510"
                }
                mac_hostname_waplist.append(DICT)

        #print(mac_hostname_waplist)
        return mac_hostname_waplist

def main():
    #   print(check_ruckus_mac("34159307F00"))
    #   host_name_checker("B98-L6A-R54")
    #   exit()

    # Main HOST Variable for API Call
    host = "192.X.X.X"
    # Ruckus Session 
    ruckus_sesh = ruckus_SZ_API(host)

    # API Version needed for log on
    api_version = ruckus_sesh.retrieve_api_version()
    print("API Version: {}".format(ruckus_sesh.api_version))

    # Generate username and password
    username = input("Username:\n")
    password = getpass("Password:\n")

    # Get the Service Ticket to include in all requests
    ruckus_sesh.service_ticket_logon(host, api_version, username, password)
    print("Service Ticket Produced: {}\n".format(ruckus_sesh.required_headers["serviceTicket"]))

    # Get ZONE ID when you specify ZONE name
    zone_name = "Zone"
    print("Retrieving ZONE ID from the ZONE NAME: {}".format(zone_name))
    ruckus_sesh.retrieve_zone_id(zone_name) 
    print("ZONE ID Retrieved: {}\n".format(ruckus_sesh.zone_id))

    # Get GROUP ID when you specify GROUP name
    group_name = "Group"
    print("Retrieving GROUP ID from the GROUP NAME: {}".format(group_name))
    ruckus_sesh.retrieve_group_id(group_name) 
    print("GROUP ID Retrieved: {}\n".format(ruckus_sesh.group_id))

    print("Time to ADD WAP/s to SZ Host")
    input("Press ENTER to CONTINUE: [\\n]")

    """
    mac = input("1. MAC Address of WAP:\n")
    host_name = input("1. Hostname of WAP:\n")
    ruckus_sesh.create_ruckus_ap(self, host_name, mac)
    """
    
    # Retrieve List of MAC Addresses and Hosts
    list_mac_hostnames = ruckus_sesh.get_list_mac_hosts("List_WAPs.xlsx")
    
    # Iterate through the list

    for machosts in list_mac_hostnames:
        try:
            print("MAC Address: {}".format(machosts['mac']))
            print("Hostname: {}".format(machosts['name']))
            print("ZONE ID: {}".format(machosts['zoneId']))
            print("GROUP ID: {}".format(machosts['groupId']))
            input("Correct?\nPress ENTER to CONTINUE: [\\n]")
            print(
            """
            Creating WAP - {}
            MAC Address : {}
            Zone ID : {}
            Group ID : {}
            """
            )
            ruckus_sesh.create_ruckus_ap(self, machosts['name'], machosts['mac'])
            
        except KeyboardInterrupt as e:
            cprint("Unable to establish session to Ruckus Smart Zone:\n  ", 'red', True)
            cprint(e,'yellow')
            exit()

if __name__ == '__main__':
    main()

